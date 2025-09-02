# excel_plot.py
from __future__ import annotations

import os
from datetime import datetime
from typing import Dict, Optional, Sequence

import numpy as np
import pandas as pd

# We use openpyxl to write an .xlsx and create a Scatter chart inside it.
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import DataPoint


DEFAULT_TENORS_YEARS: Dict[str, float] = {
    "13 semaines": 13 / 52,
    "26 semaines": 26 / 52,
    "52 semaines": 1,
    "2 ans": 2,
    "5 ans": 5,
    "10 ans": 10,
    "15 ans": 15,
    "20 ans": 20,
    "30 ans": 30,
}


def _read_and_clean(csv_path: str) -> pd.DataFrame:
    """
    Read BKAM CSV (no header) and compute time-to-maturity (years).
    Columns: [maturity, col1, rate, ref_date]
    """
    df = pd.read_csv(
        csv_path,
        header=None,
        names=["maturity", "col1", "rate", "ref_date"],
        dtype=str
    )

    # Clean numeric rate ("3,25%" -> "3.25")
    df["rate"] = (
        df["rate"]
        .astype(str)
        .str.replace(",", ".", regex=False)
        .str.replace("%", "", regex=False)
        .str.replace("\u00A0", "", regex=False)
        .str.strip()
    )
    df = df[df["rate"].str.match(r"^\d+(\.\d+)?$")]
    df["rate"] = df["rate"].astype(float)

    # Dates
    df["maturity"] = pd.to_datetime(df["maturity"], dayfirst=True, errors="coerce")
    ref_date = pd.to_datetime(df["ref_date"].iloc[0], dayfirst=True, errors="coerce")

    # Time to maturity in years
    df["ttm_years"] = (df["maturity"] - ref_date).dt.days / 365.0
    df = df.dropna(subset=["ttm_years"]).sort_values("ttm_years")
    df = df[df["ttm_years"] >= 0]

    # Final tidy columns for Excel
    out = pd.DataFrame({
        "MaturityDate": df["maturity"],
        "RefDate": ref_date,
        "TTM_Years": df["ttm_years"],
        "Rate_%": df["rate"],
    })
    return out


def _interpolate_tenors(df: pd.DataFrame, tenors_years: Dict[str, float]) -> pd.DataFrame:
    """Linear interpolation at standard tenors (in years)."""
    if df.empty or df["TTM_Years"].nunique() < 2:
        return pd.DataFrame(columns=["TenorLabel", "TenorYears", "Rate_%"])

    x = df["TTM_Years"].to_numpy()
    y = df["Rate_%"].to_numpy()

    order = np.argsort(x)
    x = x[order]
    y = y[order]

    labels = list(tenors_years.keys())
    tvals = np.array(list(tenors_years.values()), dtype=float)

    # Clip to data range to avoid extrapolation beyond ends
    tvals_clip = np.clip(tvals, x.min(), x.max())
    y_interp = np.interp(tvals_clip, x, y)

    return pd.DataFrame({
        "TenorLabel": labels,
        "TenorYears": tvals,
        "Rate_%": y_interp,
    })


from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines

def _nice_axis(min_v: float, max_v: float, target_ticks: int = 6):
    """
    Compute 'nice' axis padding and a major unit. Returns (vmin, vmax, major_unit).
    """
    if min_v == max_v:
        # Expand a tiny bit if all values are equal
        span = max(1.0, abs(min_v) * 0.1)
        min_v -= span
        max_v += span

    span = max_v - min_v
    # Candidate steps (covers most yield/time scales)
    steps = [0.1, 0.2, 0.25, 0.5, 1, 2, 2.5, 5, 10]
    # Ideal step
    ideal = span / max(2, target_ticks)
    step = min(steps, key=lambda s: abs(s - ideal))

    # Pad to multiples of step
    import math
    vmin = math.floor(min_v / step) * step
    vmax = math.ceil(max_v / step) * step
    # Recompute span in case rounding expanded range
    return vmin, vmax, step

def export_yield_curve_to_excel(
    csv_path: str,
    xlsx_target: Optional[str] = None,
    *,
    title: str = "Courbe des taux souverains (Marché secondaire)",
    include_tenors: bool = True,
    tenors_years: Optional[Dict[str, float]] = None,
    chart_position: str = "B3",
    y_as_percent: bool = False,      # NEW: format y axis as real % (divide by 100)
) -> str:
    tenors_years = tenors_years or DEFAULT_TENORS_YEARS
    data_df = _read_and_clean(csv_path)
    if data_df.empty:
        raise ValueError("No valid rows after cleaning the CSV.")

    # Optionally convert to fraction for % axis formatting
    work_df = data_df.copy()
    if y_as_percent:
        work_df["Rate_%"] = work_df["Rate_%"] / 100.0

    tenors_df = _interpolate_tenors(work_df, tenors_years) if include_tenors else pd.DataFrame()

    # Default output path
    if not xlsx_target:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_dir = os.path.dirname(csv_path) or "."
        xlsx_target = os.path.join(base_dir, f"yield_curve_{ts}.xlsx")
    os.makedirs(os.path.dirname(xlsx_target) or ".", exist_ok=True)

    # Workbook + sheets
    wb = Workbook()

    ws_data = wb.active
    ws_data.title = "Data"
    for r in dataframe_to_rows(work_df, index=False, header=True):
        ws_data.append(r)

    if include_tenors and not tenors_df.empty:
        ws_ten = wb.create_sheet("Tenors")
        for r in dataframe_to_rows(tenors_df, index=False, header=True):
            ws_ten.append(r)

    ws_chart = wb.create_sheet("Chart")

    # === Chart ===
    chart = ScatterChart()
    # Single-line title to avoid overlap on some Excel versions
    ref_date_str = work_df["RefDate"].iloc[0].strftime("%d/%m/%Y")
    chart.title = f"{title} — {ref_date_str}"
    chart.style = 2
    chart.legend.position = "r"
    chart.width = 28  # ~ wide
    chart.height = 15 # ~ tall

    chart.x_axis.title = "Échéance (années)"
    chart.y_axis.title = "Taux (%)" if not y_as_percent else "Taux"

    # Gridlines for clarity
    chart.x_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines = ChartLines()

    # Data series (raw curve)
    nrows = len(work_df) + 1
    x_ref = Reference(ws_data, min_col=3, min_row=2, max_row=nrows)  # TTM_Years
    y_ref = Reference(ws_data, min_col=4, min_row=2, max_row=nrows)  # Rate_%
    series_raw = Series(y_ref, xvalues=x_ref, title="Courbe brute")
    series_raw.marker.symbol = "circle"
    series_raw.marker.size = 7
    series_raw.graphicalProperties.line.width = 19050  # ~2.0 pt
    chart.series.append(series_raw)

    # Tenor points (optional)
    if include_tenors and not tenors_df.empty:
        nrows_t = len(tenors_df) + 1
        ws_ten = wb["Tenors"]
        x_ref_t = Reference(ws_ten, min_col=2, min_row=2, max_row=nrows_t)  # TenorYears
        y_ref_t = Reference(ws_ten, min_col=3, min_row=2, max_row=nrows_t)  # Rate_%
        series_ten = Series(y_ref_t, xvalues=x_ref_t, title="Tenors (interpolés)")
        series_ten.marker.symbol = "diamond"
        series_ten.marker.size = 8
        series_ten.graphicalProperties.line.noFill = True
        chart.series.append(series_ten)

    # === Axis formatting ===
    # X axis (years)
    xmin = float(work_df["TTM_Years"].min())
    xmax = float(work_df["TTM_Years"].max())
    xmin, xmax, xstep = _nice_axis(xmin, xmax, target_ticks=6)
    chart.x_axis.scaling.min = xmin
    chart.x_axis.scaling.max = xmax
    chart.x_axis.majorUnit = xstep
    chart.x_axis.number_format = "0.0" if (xmax - xmin) <= 3 else "0"

    # Y axis (rates)
    ymin = float(work_df["Rate_%"].min())
    ymax = float(work_df["Rate_%"].max())
    ymin, ymax, ystep = _nice_axis(ymin, ymax, target_ticks=6)
    chart.y_axis.scaling.min = ymin
    chart.y_axis.scaling.max = ymax
    chart.y_axis.majorUnit = ystep
    chart.y_axis.number_format = "0.00%" if y_as_percent else "0.00"

    # Place chart
    ws_chart.add_chart(chart, chart_position)

    wb.save(xlsx_target)
    return xlsx_target

